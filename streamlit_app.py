"""Thai Streamlit interface; calculations remain in calculation_engine.py."""
from __future__ import annotations

import json
import re
from io import BytesIO

import pandas as pd
import streamlit as st

from calculation_engine import (
    DEFAULT_INVERTERS, DEFAULT_MODULES, calculate_design, calculate_string_limits, csv_bytes,
    inverter_optimisation, make_pvsyst_export, qa_summary, recommend_inverter_options,
    recommend_string_groups,
)

st.set_page_config(page_title="Solar Rooftop Design Assistant", page_icon="☀️", layout="wide")
st.markdown("""<style>
.block-container {padding-top:1.25rem; padding-bottom:2rem;}
div[data-testid="stMetric"] {background:#f1f7f5;border:1px solid #c8ded6;padding:.55rem;border-radius:.5rem;}
div[data-testid="stDataFrame"] [role="columnheader"] {
    white-space:normal !important;
    line-height:1.15 !important;
}
.important-field-title {
    background:linear-gradient(90deg,#fff3a3 0%,#fffbe6 100%);
    border:2px solid #eab308;
    border-left:7px solid #ca8a04;
    border-radius:.45rem;
    color:#713f12;
    font-weight:800;
    padding:.42rem .7rem;
    margin-bottom:.35rem;
}
.important-field-note {
    color:#854d0e;
    font-size:.82rem;
    margin:.1rem 0 .5rem;
}
.selected-inverter-caption {
    background:linear-gradient(90deg,#dcfce7 0%,#f0fdf4 100%);
    border:2px solid #16a34a;
    border-left:7px solid #15803d;
    border-radius:.45rem;
    color:#14532d;
    padding:.55rem .75rem;
    margin-top:.45rem;
    line-height:1.35;
}
.selected-inverter-caption-title {
    font-size:.8rem;
    font-weight:800;
    text-transform:uppercase;
    letter-spacing:.02em;
}
.selected-inverter-caption-model {
    font-size:1rem;
    font-weight:800;
}
.inverter-set-highlight {
    display:flex;
    align-items:center;
    gap:.65rem;
    flex-wrap:wrap;
    background:linear-gradient(90deg,#fff3a3 0%,#fffbe6 100%);
    border:2px solid #eab308;
    border-left:7px solid #ca8a04;
    border-radius:.5rem;
    color:#713f12;
    padding:.55rem .75rem;
    margin:.55rem 0 .7rem;
}
.inverter-set-highlight-title {
    font-weight:800;
}
.inverter-set-chip {
    display:inline-block;
    background:#ffffff;
    border:2px solid #eab308;
    border-radius:.35rem;
    color:#713f12;
    font-weight:800;
    padding:.18rem .55rem;
    margin:.1rem .15rem;
}
</style>""", unsafe_allow_html=True)

ROOF_COLUMNS = [
    "roof_id", "zone", "group_id", "modules", "inverter_override", "orientation",
    "tilt_deg", "azimuth_deg", "shading", "one_way_m", "mppt_override", "input_override",
]
INVERTER_COLORS = [
    ("#dbeafe", "#1e3a8a"),  # blue
    ("#dcfce7", "#14532d"),  # green
    ("#fef3c7", "#78350f"),  # amber
    ("#fce7f3", "#831843"),  # pink
    ("#ede9fe", "#4c1d95"),  # violet
    ("#cffafe", "#164e63"),  # cyan
    ("#ffedd5", "#7c2d12"),  # orange
    ("#e2e8f0", "#1e293b"),  # slate
]


def inverter_colors(inverter_id: object) -> tuple[str, str]:
    """Return stable background/text colors for an INVxx identifier."""
    value = str(inverter_id)
    if value == "UNASSIGNED":
        return "#fee2e2", "#991b1b"
    match = re.search(r"(\d+)$", value)
    if not match:
        return "#f3f4f6", "#374151"
    return INVERTER_COLORS[(int(match.group(1)) - 1) % len(INVERTER_COLORS)]


def inverter_cell_style(value: object) -> str:
    background, foreground = inverter_colors(value)
    return (
        f"background-color:{background};color:{foreground};"
        "font-weight:700;border-left:4px solid currentColor"
    )


def default_roof_groups() -> pd.DataFrame:
    """Return a new frame so sessions never share or mutate the same object."""
    return pd.DataFrame([
         ["RF01", "Upper", "G01", 18, "AUTO", "Portrait", 10, 180, "Low", 35, "AUTO", "AUTO"],
         ["RF01", "Upper", "G02", 18, "AUTO", "Portrait", 10, 180, "Low", 40, "AUTO", "AUTO"],
         ["RF01", "Upper", "G03", 18, "AUTO", "Portrait", 10, 180, "Low", 45, "AUTO", "AUTO"],
         ["RF02", "Lower", "G04", 16, "AUTO", "Portrait", 10, 180, "Low", 55, "AUTO", "AUTO"],
         ["RF02", "Lower", "G05", 16, "AUTO", "Portrait", 10, 180, "Low", 60, "AUTO", "AUTO"],
         ["RF02", "Lower", "G06", 16, "AUTO", "Portrait", 10, 180, "Low", 65, "AUTO", "AUTO"],
    ], columns=ROOF_COLUMNS)


def _is_inverter_token(value: object) -> bool:
    normalized = str(value).strip().upper()
    return normalized == "AUTO" or bool(re.fullmatch(r"INV0*\d+", normalized))


def _is_numeric_token(value: object) -> bool:
    try:
        float(str(value).strip().replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _normalise_slot_token(value: object) -> str:
    """Keep imported MPPT/input overrides in a stable AUTO or integer form."""
    text = "" if value is None else str(value).strip()
    if not text or text.upper() == "AUTO":
        return "AUTO"
    try:
        number = float(text.replace(",", ""))
        if number.is_integer() and number > 0:
            return str(int(number))
    except (TypeError, ValueError):
        pass
    return text.upper()


def _parse_excel_clipboard_legacy(text: str) -> pd.DataFrame:
    """Legacy parser retained for reference; use parse_excel_clipboard below."""
    rows = [
        (
            [cell.strip() for cell in line.split("\t")]
            if "\t" in line
            else [cell.strip() for cell in re.split(r"\s{2,}", line.strip())]
        )
        for line in text.splitlines()
        if line.strip()
    ]
    if not rows:
        raise ValueError("ยังไม่มีข้อมูลที่ Paste")
    first_cell = rows[0][0].strip().lower().replace(" ", "_")
    if first_cell in {"roof_id", "roofid"}:
        rows = rows[1:]
    if not rows:
        raise ValueError("พบเฉพาะหัวตาราง แต่ไม่มีแถวข้อมูล")

    parsed_rows = []
    for row_number, row in enumerate(rows, start=1):
        if len(row) == 10:
            if _is_inverter_token(row[4]):
                # Input-only: roof, zone, group, modules, inverter, orientation,
                # tilt, azimuth, shading, one-way.
                values = row
            elif _is_numeric_token(row[4]) and _is_inverter_token(row[5]):
                # Compact legacy: roof, zone, group, modules, calculated kWp,
                # inverter, orientation, tilt, azimuth, shading. One-way is omitted.
                values = [row[0], row[1], row[2], row[3], row[5],
                          row[6], row[7], row[8], row[9], None]
            else:
                values = row
        elif len(row) == 11:
            # Legacy display: includes calculated kWp at index 4.
            values = [row[0], row[1], row[2], row[3], row[5],
                      row[6], row[7], row[8], row[9], row[10]]
        elif len(row) == 12:
            # Current display: includes calculated kWp and Assigned Inverter.
            values = [row[0], row[1], row[2], row[3], row[5],
                      row[7], row[8], row[9], row[10], row[11]]
        else:
            raise ValueError(
                f"แถว {row_number} มี {len(row)} คอลัมน์ "
                "(รองรับ 10, 11 หรือ 12 คอลัมน์)"
            )
        try:
            modules = int(float(values[3].replace(",", "")))
            tilt = float(values[6].replace(",", "")) if values[6] else None
            azimuth = float(values[7].replace(",", "")) if values[7] else None
            one_way = float(values[9].replace(",", "")) if values[9] else None
        except ValueError as error:
            raise ValueError(f"แถว {row_number} มีค่าตัวเลขไม่ถูกต้อง") from error
        inverter_value = values[4].upper() if values[4] else "AUTO"
        match = re.fullmatch(r"INV0*(\d+)", inverter_value)
        if match:
            inverter_value = f"INV{int(match.group(1)):02d}"
        parsed_rows.append([
            values[0], values[1], values[2], modules, inverter_value,
            values[5], tilt, azimuth, values[8], one_way,
        ])
    legacy_frame = pd.DataFrame(parsed_rows, columns=ROOF_COLUMNS[:10])
    legacy_frame["mppt_override"] = "AUTO"
    legacy_frame["input_override"] = "AUTO"
    return legacy_frame.reindex(columns=ROOF_COLUMNS)


def parse_excel_clipboard(text: str) -> pd.DataFrame:
    """Parse legacy input plus optional MPPT and String overrides.

    The last two columns in the new input format are MPPT and String.  Both
    accept AUTO or a positive integer. Existing 10/11/12-column clipboard
    formats remain supported for backward compatibility.
    """
    rows = [
        ([cell.strip() for cell in line.split("\t")]
         if "\t" in line else [cell.strip() for cell in re.split(r"\s{2,}", line.strip())])
        for line in text.splitlines()
        if line.strip()
    ]
    if not rows:
        raise ValueError("ยังไม่มีข้อมูลที่ Paste")
    first_cell = rows[0][0].strip().lower().replace(" ", "_")
    if first_cell in {"roof_id", "roofid"}:
        rows = rows[1:]
    if not rows:
        raise ValueError("พบเฉพาะหัวตาราง แต่ไม่มีแถวข้อมูล")

    parsed_rows = []
    for row_number, row in enumerate(rows, start=1):
        mppt_override, input_override = "AUTO", "AUTO"
        if len(row) == 10:
            if _is_inverter_token(row[4]):
                values = [*row, "AUTO", "AUTO"]
            elif _is_numeric_token(row[4]) and _is_inverter_token(row[5]):
                values = [row[0], row[1], row[2], row[3], row[5],
                          row[6], row[7], row[8], row[9], None, "AUTO", "AUTO"]
            else:
                values = [*row, "AUTO", "AUTO"]
        elif len(row) == 11:
            # Legacy display: calculated kWp is at index 4.
            values = [row[0], row[1], row[2], row[3], row[5],
                      row[6], row[7], row[8], row[9], row[10], "AUTO", "AUTO"]
        elif len(row) == 12:
            if _is_inverter_token(row[4]):
                # New input-only: the final two columns are MPPT and String.
                values = [row[0], row[1], row[2], row[3], row[4],
                          row[5], row[6], row[7], row[8], row[9], row[10], row[11]]
            elif _is_numeric_token(row[4]) and _is_inverter_token(row[5]) and not _is_inverter_token(row[6]):
                # Compact legacy with calculated kWp and the final two overrides.
                values = [row[0], row[1], row[2], row[3], row[5],
                          row[6], row[7], row[8], row[9], None, row[10], row[11]]
            else:
                # Current display: calculated kWp + selected and assigned Inverter.
                values = [row[0], row[1], row[2], row[3], row[5],
                          row[7], row[8], row[9], row[10], row[11], "AUTO", "AUTO"]
        elif len(row) == 13:
            # Legacy display plus MPPT and String overrides.
            values = [row[0], row[1], row[2], row[3], row[5],
                      row[6], row[7], row[8], row[9], row[10], row[11], row[12]]
        elif len(row) == 14:
            # Current display plus MPPT and String overrides.
            values = [row[0], row[1], row[2], row[3], row[5],
                      row[7], row[8], row[9], row[10], row[11], row[12], row[13]]
        else:
            raise ValueError(
                f"แถว {row_number} มี {len(row)} คอลัมน์ "
                "(รองรับ 10, 11, 12, 13 หรือ 14 คอลัมน์)"
            )
        try:
            modules = int(float(str(values[3]).replace(",", "")))
            tilt = float(str(values[6]).replace(",", "")) if values[6] else None
            azimuth = float(str(values[7]).replace(",", "")) if values[7] else None
            one_way = float(str(values[9]).replace(",", "")) if values[9] else None
        except (TypeError, ValueError) as error:
            raise ValueError(f"แถว {row_number} มีค่าตัวเลขไม่ถูกต้อง") from error
        inverter_value = str(values[4]).upper() if values[4] else "AUTO"
        match = re.fullmatch(r"INV0*(\d+)", inverter_value)
        if match:
            inverter_value = f"INV{int(match.group(1)):02d}"
        mppt_override = _normalise_slot_token(values[10])
        input_override = _normalise_slot_token(values[11])
        parsed_rows.append([
            values[0], values[1], values[2], modules, inverter_value,
            values[5], tilt, azimuth, values[8], one_way,
            mppt_override, input_override,
        ])
    return pd.DataFrame(parsed_rows, columns=ROOF_COLUMNS)


def status_style(value: object) -> str:
    value = str(value)
    if value == "PASS":
        return "background-color:#d9ead3;color:#274e13;font-weight:700"
    if value == "FAIL" or value.startswith("FAIL"):
        return "background-color:#f4cccc;color:#990000;font-weight:700"
    if value == "WARNING" or value.startswith("WARNING"):
        return "background-color:#fff2cc;color:#7f6000;font-weight:700"
    if "VERIFICATION" in value:
        return "background-color:#fff2cc;color:#7f6000;font-weight:700"
    return ""


MPPT_HIGHLIGHT_COLORS = [
    "#e8f1ff", "#e8f8ef", "#fff5db", "#fbeafa",
    "#ebefff", "#e3f7f8", "#ffede3", "#f0f2f5",
    "#f3e8ff", "#ecfccb", "#fee2e2", "#cffafe",
]
STRING_HIGHLIGHT_COLORS = [
    "#2563eb", "#16a34a", "#d97706", "#c026d3",
    "#4f46e5", "#0891b2", "#ea580c", "#475569",
    "#9333ea", "#65a30d", "#dc2626", "#0e7490",
]


def _sequence_number(value: object) -> int | None:
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def assignment_row_style(row: pd.Series) -> list[str]:
    """Highlight MPPT groups and String order without hiding status colors."""
    mppt_number = _sequence_number(row.get("MPPT"))
    string_identifier = row.get("String ID", row.get("Group ID"))
    string_number = _sequence_number(string_identifier)
    mppt_color = (
        MPPT_HIGHLIGHT_COLORS[(mppt_number - 1) % len(MPPT_HIGHLIGHT_COLORS)]
        if mppt_number else "#ffffff"
    )
    string_color = (
        STRING_HIGHLIGHT_COLORS[(string_number - 1) % len(STRING_HIGHLIGHT_COLORS)]
        if string_number else "#94a3b8"
    )
    styles = [f"background-color:{mppt_color};" for _ in row.index]
    for column in ("String ID", "Group ID", "String"):
        if column in row.index:
            styles[row.index.get_loc(column)] = (
                f"background-color:{mppt_color};color:{string_color};"
                f"font-weight:700;border-left:5px solid {string_color};"
            )
    if "MPPT" in row.index:
        styles[row.index.get_loc("MPPT")] = (
            f"background-color:{mppt_color};font-weight:800;"
            "border:2px solid #64748b;"
        )
    return styles


def assignment_warning_row_style(row: pd.Series) -> list[str]:
    """Make MPPT-current warning rows unmistakable in the assignment table."""
    if str(row.get("MPPT current", "")).upper() != "WARNING":
        return ["" for _ in row.index]
    return [
        "background-color:#fecaca;color:#991b1b;font-weight:800;"
        "border-top:2px solid #dc2626;border-bottom:2px solid #dc2626;"
        for _ in row.index
    ]


def warning_index_row_style(row: pd.Series, warning_rows: set[int]) -> list[str]:
    """Apply the same red highlight to export rows by source-row index."""
    if row.name not in warning_rows:
        return ["" for _ in row.index]
    return [
        "background-color:#fecaca;color:#991b1b;font-weight:800;"
        "border-top:2px solid #dc2626;border-bottom:2px solid #dc2626;"
        for _ in row.index
    ]


def make_string_export_frame(candidate_frame: pd.DataFrame) -> pd.DataFrame:
    """Build the final Excel table in the approved 14-column order."""
    export_frame = candidate_frame[[
        "roof_id", "zone", "group_id", "modules", "string_kwp",
        "inverter_id", "inverter_model", "mppt_no", "input_no",
        "mppt_total_modules", "tilt_deg", "azimuth_deg", "shading",
        "one_way_m",
    ]].copy()
    return export_frame.rename(columns={
        "roof_id": "Roof ID",
        "zone": "Zone",
        "group_id": "String ID",
        "modules": "Modules",
        "string_kwp": "DC (kWp)",
        "inverter_id": "Assigned Inverter",
        "inverter_model": "Inverter model",
        "mppt_no": "MPPT",
        "input_no": "String",
        "mppt_total_modules": "MPPT total modules",
        "tilt_deg": "Tilt (deg)",
        "azimuth_deg": "Azimuth (deg)",
        "shading": "Shading",
        "one_way_m": "One-way cable (m)",
    })


def make_input_assignment_export_frame(candidate_frame: pd.DataFrame) -> pd.DataFrame:
    """Build the combined 16-column table shown in the Page 1 export request."""
    export_frame = candidate_frame[[
        "roof_id", "zone", "group_id", "modules", "string_kwp",
        "inverter_override", "orientation", "tilt_deg", "azimuth_deg",
        "shading", "one_way_m", "inverter_id", "inverter_model", "mppt_no",
        "input_no", "mppt_total_modules",
    ]].copy()
    return export_frame.rename(columns={
        "roof_id": "Roof ID",
        "zone": "Zone",
        "group_id": "Group ID",
        "modules": "Modules",
        "string_kwp": "DC (kWp)",
        "inverter_override": "เลือก Inverter",
        "orientation": "Orientation",
        "tilt_deg": "Tilt (deg)",
        "azimuth_deg": "Azimuth (deg)",
        "shading": "Shading",
        "one_way_m": "One-way cable (m) optional",
        "inverter_id": "Assigned Inverter",
        "inverter_model": "Inverter model",
        "mppt_no": "MPPT",
        "input_no": "String",
        "mppt_total_modules": "MPPT total modules",
    })


def string_export_xlsx_bytes(export_frame: pd.DataFrame) -> bytes:
    """Serialize the final String/MPPT table as a formatted Excel workbook."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_frame.to_excel(writer, index=False, sheet_name="String Export")
        worksheet = writer.sheets["String Export"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="0F766E")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)
        column_indexes = {
            cell.value: cell.column - 1 for cell in worksheet[1]
        }
        for row in worksheet.iter_rows(min_row=2):
            for column_name in ("Modules", "MPPT", "String", "MPPT total modules"):
                index = column_indexes.get(column_name)
                if index is not None:
                    row[index].number_format = "0"
            for column_name in ("DC (kWp)", "Tilt (deg)", "Azimuth (deg)", "One-way cable (m)", "One-way cable (m) optional"):
                index = column_indexes.get(column_name)
                if index is not None:
                    row[index].number_format = "#,##0.000" if column_name == "DC (kWp)" else "0.00"
    return output.getvalue()


def display(frame: pd.DataFrame, status_columns: list[str]) -> None:
    if frame.empty:
        st.info("ยังไม่มีข้อมูลที่แสดงผล")
        return
    display_frame = frame.copy()
    percent_labels = {
        "voltage_drop_pct": "Voltage drop (%)",
        "power_loss_pct": "Power loss (%)",
        "dc_loss_pct": "DC loss at STC (%)",
    }
    for column, label in percent_labels.items():
        if column in display_frame.columns:
            display_frame[label] = display_frame[column].map(lambda value: "-" if pd.isna(value) else f"{float(value):.2f}%")
            display_frame = display_frame.drop(columns=[column])
    display_frame = display_frame.rename(columns={
        "tilt_deg": "Tilt (deg)",
        "azimuth_deg": "Azimuth (deg)",
        "avg_one_way_m": "Average one-way cable (m)",
        "avg_loop_m": "Average loop cable (m)",
        "input_no": "String",
        "mppt_modules_per_string": "Modules/String in MPPT",
        "mppt_string_count": "Strings/MPPT",
        "mppt_total_modules": "MPPT total modules",
        "mppt_imp_a": "MPPT Imp (A)",
        "mppt_isc_a": "MPPT Isc (A)",
        "mppt_current_status": "MPPT current status",
    })
    styler = display_frame.style
    kwp_columns = [
        column for column in display_frame.columns
        if "kwp" in str(column).lower()
    ]
    if kwp_columns:
        styler = styler.format(
            {column: "{:,.3f}" for column in kwp_columns},
            na_rep="-",
        )
    for col in status_columns:
        display_col = {
            "mppt_current_status": "MPPT current status",
        }.get(col, col)
        if display_col in display_frame.columns:
            styler = styler.map(status_style, subset=[display_col])
    st.dataframe(styler, use_container_width=True, hide_index=True)


def totals_bar(strings: pd.DataFrame, title: str = "สรุปรวม", total_ac_kw: float | None = None) -> None:
    """Show consistent project totals immediately above schedule tables."""
    if strings.empty:
        return
    total_modules = int(pd.to_numeric(strings.get("modules"), errors="coerce").fillna(0).sum())
    total_strings = int(strings["string_id"].nunique()) if "string_id" in strings else len(strings)
    total_kwp = pd.to_numeric(strings.get("string_kwp"), errors="coerce").fillna(0).sum()
    electrical_pass = strings.get("electrical_status", pd.Series(dtype=str)).eq("PASS")
    assignment_pass = strings.get("assignment_status", pd.Series(dtype=str)).eq("PASS")
    passed = int((electrical_pass & assignment_pass).sum()) if "assignment_status" in strings else int(electrical_pass.sum())
    st.caption(title)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total modules", f"{total_modules:,}")
    c2.metric("Total strings", f"{total_strings:,}")
    c3.metric("Total AC capacity", f"{total_ac_kw * 1000:,.0f} W" if total_ac_kw is not None else "-")
    c4.metric("Total DC", f"{total_kwp:,.3f} kWp")
    c5.metric("Strings PASS", f"{passed:,}/{total_strings:,}")


def init_state() -> None:
    if "module_master" not in st.session_state:
        st.session_state.module_master = DEFAULT_MODULES.copy()
    if "inverter_master" not in st.session_state:
        st.session_state.inverter_master = DEFAULT_INVERTERS.copy()
    else:
        # Refresh the former placeholder row when a browser session survives
        # a code reload after the SG350HX-20 datasheet is added.
        inverter_master = st.session_state.inverter_master.copy()
        legacy_mask = inverter_master["inverter_id"].astype(str).eq("SG350HX")
        current_ids = inverter_master["inverter_id"].astype(str).tolist()
        if legacy_mask.any() and "SG350HX-20" not in current_ids:
            replacement = DEFAULT_INVERTERS.loc[
                DEFAULT_INVERTERS["inverter_id"].eq("SG350HX-20")
            ].iloc[0].to_dict()
            for column, value in replacement.items():
                inverter_master.loc[legacy_mask, column] = value
            st.session_state.inverter_master = inverter_master
    if "roof_groups" not in st.session_state:
        st.session_state.roof_groups = default_roof_groups()
    else:
        # Non-destructive migration for sessions created before manual Inverter
        # selection was added.
        migrated_roof_groups = (
            st.session_state.roof_groups.reindex(columns=ROOF_COLUMNS)
            .copy()
            .reset_index(drop=True)
        )
        migrated_roof_groups["inverter_override"] = (
            migrated_roof_groups["inverter_override"].fillna("AUTO")
        )
        for override_column in ("mppt_override", "input_override"):
            migrated_roof_groups[override_column] = (
                migrated_roof_groups[override_column].fillna("AUTO")
            )
        st.session_state.roof_groups = migrated_roof_groups
    if "roof_editor_revision" not in st.session_state:
        st.session_state.roof_editor_revision = 0


def apply_pending_auto_layout() -> None:
    pending = st.session_state.pop("pending_auto_layout", None)
    if pending is not None:
        st.session_state.roof_groups = pending.copy()
        st.session_state.roof_editor_revision = (
            int(st.session_state.get("roof_editor_revision", 0)) + 1
        )


init_state()
apply_pending_auto_layout()
st.title("☀️ Solar Rooftop String & MPPT Design Assistant")
st.caption("ออกแบบเบื้องต้น • แยก Calculation Engine และ Excel/Web Interface • เตรียมข้อมูล PVsyst")
st.warning("ผลลัพธ์เป็นเครื่องมือช่วยออกแบบเท่านั้น ต้องยืนยันกับ datasheet ล่าสุด มาตรฐาน/การไฟฟ้า สภาพหน้างาน และวิศวกรผู้มีใบอนุญาต")

with st.sidebar:
    st.header("โครงการ")
    project_name = st.text_input("ชื่อโครงการ", "Solar Rooftop Sample")
    customer = st.text_input("ลูกค้า", "ตัวอย่าง")
    st.divider()
    st.caption("Phase 1: Master + String + MPPT + QA\n\nPhase 2: Cable + Inverter optimization + PVsyst\n\nPhase 3: String Map / Report\n\nPhase 4: Integration")

tabs = st.tabs(["1. ข้อมูลตั้งต้น", "2. Auto-layout & String", "3. MPPT & Cable", "4. QA/QC & PVsyst", "5. สูตรคำนวณ", "6. คู่มือใช้งาน", "Master Data"])
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = tabs

master_modules = st.session_state.module_master
master_inverters = st.session_state.inverter_master
with tab1:
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("แผง Solar PV")
        selected_module = st.selectbox("รุ่นแผง", master_modules["module_id"].tolist())
        module = master_modules.loc[master_modules.module_id == selected_module].iloc[0].to_dict()
        module_power = st.number_input("กำลังแผง (W) — ปรับแล้ว kWp จะเปลี่ยนทั้งระบบ", 1.0, 1000.0, float(module["pmax_w"]), 1.0)
        module["pmax_w"] = module_power
        suffix = st.text_input("Suffix / รุ่นย่อยที่ยืนยันแล้ว", "BDV")
        st.caption(f"{module['manufacturer']} | {module['model']} | {module['verification_status']}")
        st.caption(
            f"Vmp {module['vmp_v']:.2f} V | Imp {module['imp_a']:.2f} A | "
            f"Voc {module['voc_v']:.2f} V | Isc {module['isc_a']:.2f} A | "
            f"Efficiency {module.get('module_efficiency_pct', 0):.2f}%"
        )
    with col2:
        st.subheader("Inverter")
        inverter_select_options = ["AUTO", *master_inverters["inverter_id"].tolist()]
        with st.container(border=True):
            st.markdown(
                "<div class='important-field-title'>⭐ ช่องสำคัญ: รุ่น Inverter</div>"
                "<div class='important-field-note'>เลือก AUTO เพื่อให้ระบบเลือกรุ่นที่ใช้จำนวนเครื่องน้อยที่สุดและจัด String/MPPT ได้ครบ</div>",
                unsafe_allow_html=True,
            )
            selected_inv_choice = st.selectbox(
                "รุ่น Inverter",
                inverter_select_options,
                index=0,
                key="initial_inverter_choice",
                label_visibility="collapsed",
                help="AUTO = เลือกรุ่นที่ใช้จำนวน Inverter น้อยที่สุดและจัด String/MPPT ให้ครบ",
            )
        selected_inv = (
            selected_inv_choice
            if selected_inv_choice != "AUTO"
            else str(master_inverters.iloc[0]["inverter_id"])
        )
        inverter = master_inverters.loc[master_inverters.inverter_id == selected_inv].iloc[0].to_dict()
        inverter_qty_options = ["AUTO", *range(1, 51)]
        inverter_qty_choice = st.selectbox(
            "จำนวน Inverter ที่ต้องการใช้",
            inverter_qty_options,
            index=0,
            format_func=lambda value: "AUTO (จำนวนน้อยที่สุดที่รองรับ)" if value == "AUTO" else str(value),
        )
        inverter_qty_input = 1 if inverter_qty_choice == "AUTO" else int(inverter_qty_choice)
        inverter_caption = st.empty()
        inverter_caption.markdown(
            "<div class='selected-inverter-caption'>"
            "<div class='selected-inverter-caption-title'>🔌 Inverter ที่ใช้ในการคำนวณ</div>"
            f"<div class='selected-inverter-caption-model'>{inverter['manufacturer']} | {inverter['model']}</div>"
            f"<div>สถานะข้อมูล: <b>{inverter['verification_status']}</b></div>"
            "</div>",
            unsafe_allow_html=True,
        )
        if inverter["verification_status"] != "Verified":
            st.error("ห้ามออกแบบ approval: รุ่นนี้ไม่มี datasheet ที่ยืนยันแล้ว")

    with st.expander("Design Basis (คลิกเพื่อขยาย)", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            tmin = st.number_input("อุณหภูมิต่ำสุด (°C)", -30.0, 40.0, 10.0, 1.0)
            tcell_max = st.number_input("อุณหภูมิเซลล์สูงสุด (°C)", 25.0, 100.0, 70.0, 1.0)
        with c2:
            safety_factor = st.number_input("Voltage safety factor", 0.80, 1.00, 0.95, 0.01)
            max_dcac = st.number_input("DC/AC ratio สูงสุด", 0.5, 2.0, 1.40, 0.01)
        with c3:
            cable_material = st.selectbox("ตัวนำสาย DC", ["Copper", "Aluminium"])
            cable_option = st.selectbox("เบอร์สาย DC", ["4 mm²", "6 mm²", "10 mm²", "16 mm²", "กำหนดเอง"])
            cable_size = st.number_input("ขนาดสายที่กำหนดเอง (mm²)", 1.0, 500.0, 6.0, 0.5) if cable_option == "กำหนดเอง" else float(cable_option.split()[0])
        with c4:
            max_voltage_drop = st.number_input("Voltage drop สูงสุด (%)", 0.1, 10.0, 1.5, 0.1) / 100
            max_dc_loss = st.number_input("DC loss สูงสุด (%)", 0.1, 10.0, 1.5, 0.1) / 100

    auto_inverter_options = pd.DataFrame()
    if selected_inv_choice == "AUTO" or inverter_qty_choice == "AUTO":
        auto_inverter_options = recommend_inverter_options(
            module=module,
            module_power_w=module_power,
            tmin_c=tmin,
            tcell_max_c=tcell_max,
            safety_factor=safety_factor,
            max_dcac=max_dcac,
            cable_material=cable_material,
            cable_size_mm2=cable_size,
            max_voltage_drop=max_voltage_drop,
            max_dc_loss=max_dc_loss,
            strings=st.session_state.roof_groups,
            inverter_master=master_inverters,
        )
        passing_options = auto_inverter_options[
            auto_inverter_options["status"].isin(["PASS", "WARNING"])
        ].copy()
        if not passing_options.empty:
            best_option = passing_options.sort_values(
                ["recommended_qty", "dc_ac_ratio", "inverter_id"],
                na_position="last",
            ).iloc[0]
            if selected_inv_choice == "AUTO":
                selected_inv = str(best_option["inverter_id"])
                inverter = master_inverters.loc[
                    master_inverters["inverter_id"].eq(selected_inv)
                ].iloc[0].to_dict()
                inverter_caption.markdown(
                    "<div class='selected-inverter-caption'>"
                    "<div class='selected-inverter-caption-title'>🔌 Inverter ที่ใช้ในการคำนวณ</div>"
                    f"<div class='selected-inverter-caption-model'>{inverter['manufacturer']} | {inverter['model']}</div>"
                    f"<div>สถานะข้อมูล: <b>{inverter['verification_status']}</b></div>"
                    "</div>",
                    unsafe_allow_html=True,
                )
            chosen_options = auto_inverter_options[
                auto_inverter_options["inverter_id"].eq(selected_inv)
                & auto_inverter_options["status"].isin(["PASS", "WARNING"])
            ]
            if chosen_options.empty:
                st.error(
                    f"รุ่น {inverter['model']} ยังจัด String/MPPT ตามเกณฑ์ไม่ได้ "
                    "กรุณาเลือก Inverter เป็น AUTO หรือแก้จำนวน String"
                )
            else:
                if inverter_qty_choice == "AUTO":
                    inverter_qty_input = int(chosen_options.iloc[0]["recommended_qty"])
                message = (
                    f"AUTO แนะนำ {inverter['model']} จำนวน {inverter_qty_input} เครื่อง "
                    "(เลือกจำนวน Inverter ต่ำสุดที่จัด String/MPPT ได้ครบ)"
                )
                if str(chosen_options.iloc[0]["status"]) == "WARNING":
                    st.warning(message + " — มี 1 String เป็นเลขคี่ ระบบนำเศษลง String เดียว")
                else:
                    st.success(message)
        else:
            st.error(
                "AUTO ยังหาแผนที่ผ่านไม่ได้ — ตรวจแรงดัน String, จำนวนแผง "
                "และให้มี String เลขคี่ได้ไม่เกิน 1 String"
            )
        if selected_inv_choice == "AUTO":
            st.caption("ตารางเปรียบเทียบ AUTO: เลือกแถว PASS/WARNING ที่มีจำนวน Inverter น้อยที่สุด")
            display(auto_inverter_options, ["status"])

    st.subheader("Roof layout / Candidate strings")
    if selected_inv_choice == "AUTO":
        st.info(
            f"Inverter ที่ AUTO เลือกให้ตารางนี้: {inverter['model']} "
            f"จำนวน {inverter_qty_input} เครื่อง — ตารางด้านล่างจะแสดงชุด Inverter, MPPT และ String ที่จัดให้"
        )
    st.markdown("<div style='background:#fff2cc;border-left:5px solid #d6b656;padding:10px;border-radius:4px'>🟨 <b>ช่องที่ต้องกรอก:</b> Roof ID, Zone, Group ID, จำนวนแผง, Orientation, Tilt, Azimuth และ Shading. ปกติจำนวนแผงต่อ String เป็นเลขคู่; หากยอดรวมเป็นเลขคี่ ระบบจะนำเศษลง 1 String และแสดง WARNING. String ต่างกันไม่เกิน 2 แผง. คอลัมน์ <b>เลือก Inverter</b> ใช้ AUTO หรือระบุ INVxx ราย String ได้ • หลัง copy/paste จาก Excel ให้กด <b>บันทึกข้อมูลและคำนวณใหม่</b> • One-way cable เว้นว่างเพื่อกรอกภายหลังได้</div>", unsafe_allow_html=True)
    st.caption("กรอกจาก drone, DWG หรือ survey • one-way cable คือระยะจริงขาเดียว • กด Submit ก่อนเปลี่ยนรุ่น/จำนวน Inverter")
    if st.session_state.pop("roof_saved_notice", False):
        st.success("บันทึกข้อมูลตารางแล้ว และคำนวณ kWp / Inverter Set ใหม่เรียบร้อย")
    duplicate_groups = st.session_state.roof_groups.duplicated(
        subset=["roof_id", "zone", "group_id"], keep=False
    )
    if duplicate_groups.any():
        duplicate_names = (
            st.session_state.roof_groups.loc[duplicate_groups, "group_id"]
            .dropna().astype(str).unique().tolist()
        )
        st.warning(
            "พบ Group ID ซ้ำใน Roof/Zone เดียวกัน: "
            + ", ".join(duplicate_names[:10])
            + (" ..." if len(duplicate_names) > 10 else "")
            + " — ควรแก้ให้ไม่ซ้ำเพื่อให้ติดตาม String ได้ถูกต้อง"
        )
    inverter_options = [
        "AUTO", *[f"INV{number:02d}" for number in range(1, inverter_qty_input + 1)]
    ]
    with st.expander("📋 Paste หลายแถวจาก Excel (รองรับ 10/11/12/13/14 คอลัมน์)"):
        with st.form("excel_clipboard_form", clear_on_submit=False):
            st.caption(
                "รองรับ 3 รูปแบบ: 10 คอลัมน์ input-only, 10 คอลัมน์แบบมี DC (kWp) "
                "แต่ไม่ต้องใส่ One-way cable และ 11/12 คอลัมน์จากตารางเดิม "
                "โดยรูปแบบ 10 คอลัมน์แบบมี DC ใช้ลำดับ: Roof ID, Zone, Group ID, "
                "Modules, DC (kWp), เลือก Inverter, Orientation, Tilt, Azimuth, Shading"
            )
            st.caption("รูปแบบใหม่รองรับ MPPT และ String ต่อท้ายข้อมูล โดยกรอก AUTO หรือเลขช่องที่ต้องการ")
            clipboard_text = st.text_area(
                "วางข้อมูลจาก Excel ที่นี่",
                key="excel_clipboard_text",
                height=150,
                placeholder="AUTO  Auto  G01  20  14.5  AUTO  TBC  0  0  TBC  AUTO  AUTO",
            )
            paste_mode = st.radio(
                "วิธีนำเข้า",
                ["เพิ่มต่อท้าย", "แทนที่ทั้งหมด"],
                horizontal=True,
            )
            excel_paste_submitted = st.form_submit_button(
                "นำเข้าข้อมูล Excel",
                type="primary",
                use_container_width=True,
            )
        if excel_paste_submitted:
            try:
                imported_roof_groups = parse_excel_clipboard(clipboard_text)
                invalid_imported_inverters = ~imported_roof_groups[
                    "inverter_override"
                ].isin(inverter_options)
                if invalid_imported_inverters.any():
                    invalid_values = imported_roof_groups.loc[
                        invalid_imported_inverters, "inverter_override"
                    ].unique().tolist()
                    raise ValueError(
                        "Inverter ไม่อยู่ในจำนวนเครื่องที่เลือก: "
                        + ", ".join(map(str, invalid_values))
                    )
                if paste_mode == "แทนที่ทั้งหมด":
                    st.session_state.roof_groups = imported_roof_groups
                else:
                    st.session_state.roof_groups = pd.concat(
                        [st.session_state.roof_groups, imported_roof_groups],
                        ignore_index=True,
                    )
                st.session_state.roof_editor_revision += 1
                st.session_state.roof_saved_notice = True
                st.rerun()
            except ValueError as error:
                st.error(str(error))

    # Calculate a live preview from the currently persisted editor rows.  The two
    # derived columns are shown in the same grid but disabled to prevent manual edits.
    candidate_preview_design = calculate_design(
        module=module, inverter=inverter, module_power_w=module_power,
        tmin_c=tmin, tcell_max_c=tcell_max, safety_factor=safety_factor,
        inverter_qty=inverter_qty_input, max_dcac=max_dcac,
        cable_material=cable_material, cable_size_mm2=cable_size,
        max_voltage_drop=max_voltage_drop, max_dc_loss=max_dc_loss,
        strings=st.session_state.roof_groups,
    )
    candidate_editor_frame = st.session_state.roof_groups.copy().reset_index(drop=True)
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("modules") + 1,
        "string_kwp",
        pd.to_numeric(candidate_editor_frame["modules"], errors="coerce")
        * module_power / 1000,
    )
    inverter_by_row = {}
    inverter_model_by_row = {}
    mppt_by_row = {}
    input_by_row = {}
    mppt_total_modules_by_row = {}
    assignment_status_by_row = {}
    mppt_current_status_by_row = {}
    mppt_current_warning_keys = set()
    if not candidate_preview_design["assignments"].empty:
        assignment_lookup = candidate_preview_design["assignments"].drop_duplicates(
            "source_row"
        ).set_index("source_row")
        inverter_by_row = assignment_lookup["inverter_id"].to_dict()
        inverter_model_by_row = {
            row_no: (
                inverter["model"]
                if str(inverter_id) != "UNASSIGNED"
                else f"{inverter['model']} (ยังไม่จัด)"
            )
            for row_no, inverter_id in inverter_by_row.items()
        }
        mppt_by_row = assignment_lookup["mppt_no"].to_dict()
        input_by_row = assignment_lookup["input_no"].to_dict()
        mppt_total_modules_by_row = assignment_lookup[
            "mppt_total_modules"
        ].to_dict()
        assignment_status_by_row = assignment_lookup[
            "assignment_status"
        ].to_dict()
        mppt_current_status_by_row = (
            assignment_lookup["mppt_current_status"].to_dict()
            if "mppt_current_status" in assignment_lookup.columns
            else {}
        )
        mppt_current_warning_keys = {
            (str(inverter_by_row.get(row_no, "")), str(mppt_by_row.get(row_no, "")))
            for row_no, status in mppt_current_status_by_row.items()
            if str(status).upper() == "WARNING"
        }
    current_warning_rows = {
        row_no
        for row_no in candidate_editor_frame.index
        if (
            str(inverter_by_row.get(row_no, "")),
            str(mppt_by_row.get(row_no, "")),
        ) in mppt_current_warning_keys
        or str(assignment_status_by_row.get(row_no, "")).upper() == "WARNING"
    }
    invalid_overrides = ~candidate_editor_frame["inverter_override"].astype(str).isin(
        inverter_options
    )
    if invalid_overrides.any():
        st.warning(
            "ค่าเลือก Inverter บางแถวเกินจำนวนเครื่องปัจจุบัน "
            "จึงแสดงเป็น AUTO กรุณาตรวจและกดบันทึกอีกครั้ง"
        )
        candidate_editor_frame.loc[invalid_overrides, "inverter_override"] = "AUTO"
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("inverter_override") + 1,
        "inverter_id",
        [inverter_by_row.get(row_no, "-") for row_no in candidate_editor_frame.index],
    )
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("inverter_id") + 1,
        "inverter_model",
        [
            inverter_model_by_row.get(row_no, "-")
            for row_no in candidate_editor_frame.index
        ],
    )
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("inverter_model") + 1,
        "mppt_no",
        [mppt_by_row.get(row_no, pd.NA) for row_no in candidate_editor_frame.index],
    )
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("mppt_no") + 1,
        "input_no",
        [input_by_row.get(row_no, pd.NA) for row_no in candidate_editor_frame.index],
    )
    candidate_editor_frame.insert(
        candidate_editor_frame.columns.get_loc("input_no") + 1,
        "mppt_total_modules",
        [
            mppt_total_modules_by_row.get(row_no, pd.NA)
            for row_no in candidate_editor_frame.index
        ],
    )

    total_candidate_modules = int(
        pd.to_numeric(candidate_editor_frame["modules"], errors="coerce")
        .fillna(0).sum()
    )
    total_candidate_kwp = float(
        pd.to_numeric(candidate_editor_frame["string_kwp"], errors="coerce")
        .fillna(0).sum()
    )
    summary_cols = st.columns(4)
    summary_cols[0].metric("Total modules", f"{total_candidate_modules:,}")
    summary_cols[1].metric("Total DC", f"{total_candidate_kwp:,.3f} kWp")
    summary_cols[2].metric("Inverter sets", f"{inverter_qty_input:,}")
    actual_dcac = candidate_preview_design.get("actual_dcac_ratio")
    summary_cols[3].metric(
        "Project DC/AC ratio",
        f"{actual_dcac:.3f}" if actual_dcac is not None else "-",
    )
    st.markdown("**จำนวน String แยกตาม Inverter**")
    inverter_cards = []
    for _, inverter_row in candidate_preview_design["inverter_summary"].iterrows():
        inverter_id = str(inverter_row["inverter_id"])
        background, foreground = inverter_colors(inverter_id)
        inverter_cards.append(
            "<div style='min-width:170px;flex:1;padding:10px 14px;"
            f"background:{background};color:{foreground};"
            "border-radius:8px;border-left:6px solid currentColor'>"
            f"<div style='font-weight:800;font-size:1.05rem'>{inverter_id}</div>"
            f"<div style='font-size:1.25rem;font-weight:800'>"
            f"{int(inverter_row['assigned_strings']):,} Strings</div>"
            f"<div>{int(inverter_row['assigned_modules']):,} modules · "
            f"{float(inverter_row['assigned_dc_kwp']):,.3f} kWp</div>"
            "</div>"
        )
    st.markdown(
        "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:4px 0 14px'>"
        + "".join(inverter_cards)
        + "</div>",
        unsafe_allow_html=True,
    )
    if candidate_preview_design.get("mppt_current_warnings"):
        st.error(
            "WARNING: พบช่อง MPPT ที่กระแสรวมเกินข้อจำกัด: "
            + ", ".join(candidate_preview_design["mppt_current_warnings"])
            + " — ตรวจ datasheet หรือปรับ MPPT/String override ก่อนใช้งาน"
        )
        st.caption("แถวที่มีสถานะ WARNING จะแสดงเป็นพื้นสีแดงในตารางผลลัพธ์ด้านล่าง")

    # A design-dependent key forces calculated kWp/INV columns to refresh when
    # the module, inverter model, module power or inverter quantity is changed.
    roof_editor_key = (
        f"roof_editor_v5_{selected_module}_{selected_inv}_"
        f"{module_power:g}W_{inverter_qty_input}INV_"
        f"R{st.session_state.roof_editor_revision}"
    )
    with st.form("roof_candidate_form", clear_on_submit=False):
        roof_submitted_top = st.form_submit_button(
            "✅ บันทึกและคำนวณข้อมูลที่ Paste",
            type="primary",
            use_container_width=True,
        )
        st.caption(
            "ตารางกรอกข้อมูลมี 11 คอลัมน์ตรงกับ Excel เดิม สามารถ Paste ตรงได้ "
            "• หากรูปแบบคอลัมน์ต่างออกไป ให้ใช้ช่อง Paste ด้านบน"
        )
        # Keep the editable grid at the exact 11-column legacy Excel order.
        # Assigned Inverter is rendered in the read-only result table below;
        # placing it inside this grid would shift pasted Excel cells.
        candidate_input_frame = candidate_editor_frame.drop(
            columns=[
                "inverter_id", "inverter_model", "mppt_no", "input_no",
                "mppt_total_modules",
            ]
        )
        candidate_editor_styler = candidate_input_frame.style.map(
            inverter_cell_style,
            subset=["inverter_override"],
        )
        edited_candidate_frame = st.data_editor(
            candidate_editor_styler, num_rows="dynamic", use_container_width=True,
            key=roof_editor_key, disabled=["string_kwp"],
            column_config={
                "roof_id": st.column_config.TextColumn(
                    "🟨 Roof ID *", required=True, width="small"
                ),
                "zone": st.column_config.TextColumn(
                    "🟨 Zone *", required=True, width="small"
                ),
                "group_id": st.column_config.TextColumn(
                    "🟨 Group ID *", required=True, width="small"
                ),
                "modules": st.column_config.NumberColumn(
                    "🟨 จำนวนแผง\n(คู่ / เศษ 1 String)", min_value=1, step=2,
                    required=True, width="small",
                    help="ปกติให้เป็นจำนวนคู่; ถ้ายอดรวมเป็นเลขคี่ อนุญาตเศษเป็นเลขคี่ได้ 1 String และต่างกันไม่เกิน 2 แผง",
                ),
                "string_kwp": st.column_config.NumberColumn(
                    "กำลัง DC\n(kWp)", format="%.3f",
                    disabled=True, width="small"
                ),
                "inverter_override": st.column_config.SelectboxColumn(
                    "เลือก\nInverter",
                    options=inverter_options,
                    required=True,
                    width="small",
                    help="AUTO = โปรแกรมแบ่งกลุ่มให้ หรือเลือก INVxx เพื่อบังคับ String นี้",
                ),
                "orientation": st.column_config.TextColumn(
                    "🟨 Orientation *", required=True, width="small"
                ),
                "tilt_deg": st.column_config.NumberColumn(
                    "🟨 Tilt\n(deg)", min_value=0, max_value=90,
                    required=True, width="small"
                ),
                "azimuth_deg": st.column_config.NumberColumn(
                    "🟨 Azimuth\n(deg)", min_value=-180, max_value=360,
                    required=True, width="small"
                ),
                "shading": st.column_config.TextColumn(
                    "🟨 Shading *", required=True, width="small"
                ),
                "one_way_m": st.column_config.NumberColumn(
                    "One-way cable\n(m) optional", min_value=0.0,
                    width="medium",
                    help="เว้นว่างได้ โปรแกรมจะคำนวณ String ก่อนและแจ้งเตือนส่วนสาย DC",
                ),
                "mppt_override": st.column_config.SelectboxColumn(
                    "MPPT\n(AUTO/ระบุ)",
                    options=["AUTO", *[str(i) for i in range(1, int(inverter["mppt_qty"]) + 1)]],
                    required=True,
                    width="small",
                    help="AUTO = ให้ระบบจัด MPPT; ระบุเลขเพื่อบังคับ String นี้ไปยัง MPPT ที่เลือก",
                ),
                "input_override": st.column_config.SelectboxColumn(
                    "String\n(AUTO/ระบุ)",
                    options=["AUTO", *[str(i) for i in range(1, int(inverter["inputs_per_mppt"]) + 1)]],
                    required=True,
                    width="small",
                    help="AUTO = ให้ระบบเลือกช่อง String; ระบุเลขเพื่อบังคับ input ลำดับที่เลือกภายใน MPPT",
                ),
            })
        roof_submitted_bottom = st.form_submit_button(
            "✅ บันทึกข้อมูลและคำนวณใหม่",
            type="primary",
            use_container_width=True,
        )
        roof_submitted = roof_submitted_top or roof_submitted_bottom
    if roof_submitted:
        edited_roof_groups = (
            edited_candidate_frame.reindex(columns=ROOF_COLUMNS)
            .copy()
            .reset_index(drop=True)
        )
        st.session_state.roof_groups = edited_roof_groups
        st.session_state.roof_editor_revision += 1
        st.session_state.roof_saved_notice = True
        st.rerun()

    st.markdown("**ผลการจัด Inverter ราย String**")
    assignment_preview = candidate_editor_frame[[
        "roof_id", "group_id", "modules", "string_kwp",
        "inverter_override", "inverter_id", "inverter_model", "mppt_no", "input_no",
        "mppt_total_modules",
    ]].rename(columns={
        "roof_id": "Roof ID",
        "group_id": "String ID",
        "modules": "Modules",
        "string_kwp": "DC (kWp)",
        "inverter_override": "เลือก Inverter",
        "inverter_id": "Assigned Inverter",
        "inverter_model": "Inverter model",
        "mppt_no": "MPPT",
        "input_no": "String",
        "mppt_total_modules": "MPPT total modules",
    })
    assignment_preview["MPPT current"] = [
        (
            "WARNING"
            if row_no in current_warning_rows
            else str(assignment_status_by_row.get(row_no, "-"))
        )
        for row_no in assignment_preview.index
    ]
    assignment_preview_styler = (
        assignment_preview.style
        .format({"DC (kWp)": "{:,.3f}"}, na_rep="-")
        .apply(assignment_row_style, axis=1)
        .map(inverter_cell_style, subset=["Assigned Inverter"])
        .apply(assignment_warning_row_style, axis=1)
    )
    st.dataframe(
        assignment_preview_styler,
        width="stretch",
        hide_index=True,
        column_config={
            "Roof ID": st.column_config.TextColumn(width="small"),
            "String ID": st.column_config.TextColumn(width="small"),
            "Modules": st.column_config.NumberColumn(width="small"),
            "DC (kWp)": st.column_config.NumberColumn(width="small"),
            "เลือก Inverter": st.column_config.TextColumn(width="small"),
            "Assigned Inverter": st.column_config.TextColumn(width="small"),
            "Inverter model": st.column_config.TextColumn(width="medium"),
            "MPPT": st.column_config.NumberColumn(width="small"),
            "String": st.column_config.NumberColumn(width="small"),
            "MPPT total modules": st.column_config.NumberColumn(width="small"),
            "MPPT current": st.column_config.TextColumn(width="small"),
        },
    )

    st.subheader("ตาราง Export Excel — String / MPPT")
    st.caption(
        "จัดลำดับคอลัมน์ตามแบบ Export: Roof ID, Zone, String ID, Modules, DC (kWp), "
        "Assigned Inverter, Inverter model, MPPT, String, MPPT total modules, "
        "Tilt, Azimuth, Shading และ One-way cable"
    )
    string_export_frame = make_string_export_frame(candidate_editor_frame)
    string_export_styler = (
        string_export_frame.style
        .format({"DC (kWp)": "{:,.3f}"}, na_rep="-")
        .apply(assignment_row_style, axis=1)
        .map(inverter_cell_style, subset=["Assigned Inverter"])
        .apply(
            lambda row: warning_index_row_style(row, current_warning_rows),
            axis=1,
        )
    )
    st.dataframe(string_export_styler, width="stretch", hide_index=True)
    st.download_button(
        "⬇️ Export ตาราง String / MPPT เป็น Excel",
        data=string_export_xlsx_bytes(string_export_frame),
        file_name="solar_string_mppt_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.subheader("ตาราง Export Excel — Input + Assignment (16 คอลัมน์)")
    st.caption(
        "ตารางรวมตามแบบล่าสุด: 1–11 คือข้อมูลจากตารางกรอก และ 12–16 คือผล Assigned Inverter / MPPT "
        "สำหรับใช้ส่งต่อ Excel"
    )
    input_assignment_export_frame = make_input_assignment_export_frame(candidate_editor_frame)
    input_assignment_export_styler = (
        input_assignment_export_frame.style
        .format({"DC (kWp)": "{:,.3f}"}, na_rep="-")
        .apply(assignment_row_style, axis=1)
        .map(inverter_cell_style, subset=["Assigned Inverter"])
        .apply(
            lambda row: warning_index_row_style(row, current_warning_rows),
            axis=1,
        )
    )
    st.dataframe(input_assignment_export_styler, width="stretch", hide_index=True)
    st.download_button(
        "⬇️ Export ตาราง Input + Assignment เป็น Excel (16 คอลัมน์)",
        data=string_export_xlsx_bytes(input_assignment_export_frame),
        file_name="solar_string_input_assignment_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

design = calculate_design(module=module, inverter=inverter, module_power_w=module_power, tmin_c=tmin,
                          tcell_max_c=tcell_max, safety_factor=safety_factor, inverter_qty=inverter_qty_input,
                          max_dcac=max_dcac, cable_material=cable_material, cable_size_mm2=cable_size,
                          max_voltage_drop=max_voltage_drop, max_dc_loss=max_dc_loss, strings=st.session_state.roof_groups)
for warning in design.get("input_warnings", []):
    if "ช่อง MPPT ที่กระแสรวมเกินข้อจำกัด" in warning:
        continue
    st.warning(warning)

with tab2:
    st.subheader("แนะนำการจัด String จากจำนวนแผงรวม")
    with st.container(border=True):
        st.markdown(
            "<div class='important-field-title'>⭐ ช่องสำคัญ: จำนวนแผงรวมที่มี</div>"
            "<div class='important-field-note'>Rapid Shutdown / Optimizer อัตรา 2:1 — ควรเป็นจำนวนคู่; ถ้าเป็นเลขคี่จะลงเศษใน 1 String และแสดง WARNING</div>",
            unsafe_allow_html=True,
        )
        total_modules = st.number_input(
            "จำนวนแผงรวมที่มี",
            min_value=1,
            max_value=100000,
            value=int(pd.to_numeric(st.session_state.roof_groups.modules, errors="coerce").fillna(0).sum()),
            step=2,
            key="auto_layout_total_modules",
            label_visibility="collapsed",
        )
    if total_modules % 2:
        st.warning(
            f"จำนวนแผงรวม {total_modules:,} เป็นเลขคี่ — ควรปรับเป็นเลขคู่เนื่องจาก Rapid Shutdown / Optimizer ใช้อัตรา 2:1; "
            "หากคำนวณต่อ ระบบจะนำเศษเหลือลง 1 String และคง WARNING ไว้"
        )
    st.caption("เกณฑ์ Engine: String ต่างกันไม่เกิน 2 แผง; ยอดรวมเลขคู่จะจัดทุก String เป็นเลขคู่, ยอดรวมเลขคี่จะมีเลขคี่ได้ 1 String")
    dc_max_values = sorted(
        pd.to_numeric(master_inverters["dc_max_v"], errors="coerce")
        .dropna()
        .astype(float)
        .unique()
        .tolist()
    )
    dc_max_scope_labels = [f"DC max V = {int(value):,} V" for value in dc_max_values]
    auto_layout_inverter_options = [
        "AUTO (อ้างอิง Inverter จากหน้า 1)",
        *dc_max_scope_labels,
    ]
    with st.container(border=True):
        st.markdown(
            "<div class='important-field-title'>⭐ ช่องสำคัญ: Scope แรงดัน DC max V</div>"
            "<div class='important-field-note'>เลือก AUTO เพื่ออ้างอิง Inverter จากหน้า 1 หรือเลือกค่า DC max V ที่มีในตาราง Master Inverter</div>",
            unsafe_allow_html=True,
        )
        auto_layout_inverter_choice = st.selectbox(
            "เลือก Scope แรงดัน DC max V จากตาราง Inverter",
            auto_layout_inverter_options,
            index=0,
            key="auto_layout_dc_scope",
            label_visibility="collapsed",
            help="DC max V ใช้กำหนดจำนวนแผงสูงสุดต่อ String; ระบบจะตรวจ MPPT min/startup และ safety factor ร่วมด้วย",
        )
    if auto_layout_inverter_choice.startswith("AUTO"):
        auto_layout_inverter = inverter
        auto_layout_comment = (
            f"AUTO อ้างอิง {inverter['model']} จากหน้า 1.ข้อมูลตั้งต้น — "
            "หากต้องการเปลี่ยนรุ่น ให้กลับไปเลือกที่หน้า 1"
        )
    else:
        selected_dc_max_v = float(dc_max_values[dc_max_scope_labels.index(auto_layout_inverter_choice)])
        auto_layout_inverter = inverter.copy()
        auto_layout_inverter["dc_max_v"] = selected_dc_max_v
        auto_layout_comment = (
            f"เลือก Scope เฉพาะ DC max V = {selected_dc_max_v:,.0f} V จากตาราง Inverter — "
            f"ช่วง MPPT/startup ยังอ้างอิง {inverter['model']} จากหน้า 1"
        )
    auto_layout_limits = calculate_string_limits(
        module, auto_layout_inverter, tmin, tcell_max, safety_factor
    )
    if auto_layout_limits:
        st.caption(
            f"{auto_layout_comment} | DC max V = {float(auto_layout_inverter['dc_max_v']):,.0f} V | "
            f"Nmin MPPT = {auto_layout_limits['nmin_mppt']} | Nmax design = {auto_layout_limits['nmax_design']} แผง — "
            "DC max V เป็นเพดานแรงดันฝั่ง PV ไม่ใช่กำลัง AC ของ Inverter"
        )
    if design.get("critical_missing"):
        st.error("ต้องกรอก datasheet inverter ให้ครบก่อนสร้างคำแนะนำ")
    elif not auto_layout_limits:
        st.error("Inverter กลุ่มนี้ยังไม่มีค่า DC max V / MPPT ที่ยืนยันแล้ว")
    else:
        auto_groups = recommend_string_groups(total_modules, auto_layout_limits, module_power)
        st.caption("คำแนะนำเริ่มจาก string ที่ใกล้เคียงกันและอยู่ในช่วงแรงดันที่ผ่าน จากนั้นจึงนำไปแยก MPPT")
        if not auto_groups.empty:
            a1, a2, a3 = st.columns(3)
            a1.metric("Auto-layout total modules", f"{int(auto_groups['modules'].sum()):,}")
            a2.metric("Auto-layout total strings", f"{len(auto_groups):,}")
            a3.metric("Auto-layout total DC", f"{auto_groups['string_kwp'].sum():,.3f} kWp")
        display(auto_groups, [])
        if st.button("ใช้ Auto-layout แทน Candidate strings") and not auto_groups.empty and auto_groups.modules.min() > 0:
            st.session_state.pending_auto_layout = pd.DataFrame([
                ["AUTO", "Auto", f"G{i+1:02d}", int(row.modules), "AUTO",
                 "TBC", 0, 0, "TBC", None, "AUTO", "AUTO"]
                for i, row in auto_groups.iterrows()
            ], columns=ROOF_COLUMNS)
            st.rerun()
        total_dc_kwp = total_modules * module_power / 1000
        st.subheader("Inverter quantity & DC/AC optimization")
        optimisation = inverter_optimisation(total_dc_kwp, max_dcac, master_inverters)
        display(optimisation, ["status"])
        st.caption("WARNING ของ DC/AC ratio ต่ำ (<0.80) ไม่ได้แปลว่าออกแบบไม่ได้ แต่ควรทบทวนจำนวน inverter, redundancy และเป้าหมายการผลิต")
        st.subheader("ผลตรวจ Candidate strings ปัจจุบัน")
        metrics = st.columns(4)
        metrics[0].metric("Nmin MPPT", f"{auto_layout_limits['nmin_mppt']} modules")
        metrics[1].metric("Nmax design", f"{auto_layout_limits['nmax_design']} modules")
        metrics[2].metric("Nmax absolute", f"{auto_layout_limits['nmax_absolute']} modules")
        metrics[3].metric("Total DC", f"{design['strings']['string_kwp'].sum():,.3f} kWp")
        totals_bar(design["strings"], "ยอดรวม Candidate Strings", design.get("total_ac_kw"))
        display(design["strings"], ["electrical_status"])

with tab3:
    st.subheader("Inverter Design Sets")
    st.caption("กระจายกำลัง DC ให้สมดุลระหว่าง Inverter ก่อน แล้วจึงจัด MPPT ภายในแต่ละชุด")
    display(design["inverter_summary"], ["status"])
    if not design["assignments"].empty:
        inverter_sets = design["inverter_summary"]["inverter_id"].tolist()
        inverter_set_chips = "".join(
            f"<span class='inverter-set-chip'>{inverter_id}</span>"
            for inverter_id in inverter_sets
        )
        st.markdown(
            "<div class='inverter-set-highlight'>"
            "<span class='inverter-set-highlight-title'>⭐ Inverter Design Sets ที่กำลังตรวจสอบ:</span>"
            f"<span>{inverter_set_chips}</span>"
            "</div>",
            unsafe_allow_html=True,
        )
        set_tabs = st.tabs(inverter_sets)
        for set_tab, inverter_id in zip(set_tabs, inverter_sets):
            with set_tab:
                set_assignments = design["assignments"][
                    design["assignments"]["inverter_id"] == inverter_id
                ]
                set_summary = design["inverter_summary"][
                    design["inverter_summary"]["inverter_id"] == inverter_id
                ].iloc[0]
                set_ac_kw = (
                    float(set_summary["rated_ac_kw"])
                    if pd.notna(set_summary["rated_ac_kw"]) else None
                )
                totals_bar(
                    set_assignments,
                    f"สรุปชุด {inverter_id}",
                    set_ac_kw,
                )
                display(set_assignments, ["assignment_status", "mppt_current_status", "electrical_status"])
    st.subheader("DC Cable Calculation")
    st.caption("Voltage drop และ Power loss แสดงเป็น % • เปลี่ยนเบอร์สายในหน้า ข้อมูลตั้งต้น แล้วผลจะคำนวณใหม่")
    totals_bar(design["cables"].merge(design["strings"][["string_id", "modules", "string_kwp", "electrical_status"]], on="string_id", how="left") if not design["cables"].empty else pd.DataFrame(), "ยอดรวม String ที่ตรวจสาย DC", design.get("total_ac_kw"))
    display(design["cables"], ["cable_status"])
    if not design["assignments"].empty:
        st.bar_chart(design["assignments"].set_index("string_id")["one_way_m"])

with tab4:
    qa = qa_summary(design, module, inverter, suffix)
    st.subheader("QA/QC และคำแนะนำ")
    p1, p2, p3 = st.columns(3)
    p1.metric("PASS", int((qa.result == "PASS").sum()))
    p2.metric("WARNING", int((qa.result == "WARNING").sum()))
    p3.metric("FAIL", int((qa.result == "FAIL").sum()))
    display(qa, ["result"])
    st.subheader("PVsyst Export Preparation")
    pvsyst = make_pvsyst_export(project_name, module, inverter, design)
    if not pvsyst.empty:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("PVsyst total modules", f"{int(pvsyst['total_modules'].sum()):,}")
        c2.metric("PVsyst sub-arrays", f"{len(pvsyst):,}")
        c3.metric("PVsyst AC capacity", f"{design.get('total_ac_kw', 0) * 1000:,.0f} W")
        c4.metric("PVsyst installed DC", f"{pvsyst['installed_dc_kwp'].sum():,.3f} kWp")
    display(pvsyst, ["electrical_status", "data_status"])
    st.download_button("ดาวน์โหลด PVsyst preparation CSV", csv_bytes(pvsyst), "pvsyst_preparation.csv", "text/csv")
    package = {"project": project_name, "customer": customer, "module": module, "inverter": inverter,
               "limits": design.get("limits", {}), "strings": design["strings"].to_dict("records"),
               "inverter_sets": design["inverter_summary"].to_dict("records"),
               "assignments": design["assignments"].to_dict("records")}
    st.download_button("ดาวน์โหลด Design Package (JSON)", json.dumps(package, ensure_ascii=False, indent=2, default=str).encode(), "solar_design_package.json", "application/json")

with tab5:
    st.subheader("สูตร Calculation Engine")
    st.markdown("""
| ผลลัพธ์ | สูตร | เกณฑ์ |
|---|---|---|
| Voc ที่อากาศเย็น | `Voc_cold = Voc_STC × [1 + abs(βVoc) × (25 − Tmin)]` | String Voc cold ≤ Inverter max DC V |
| Vmp ที่อุณหภูมิร้อน | `Vmp_hot = Vmp_STC × [1 + βVmp × (Tcell,max − 25)]` | อยู่ในช่วง Start-up / MPPT min–max |
| Nmax absolute | `FLOOR(Max DC voltage / Voc_cold)` | ขีดจำกัดทางกายภาพ |
| Nmax design | `FLOOR(Max DC voltage × safety factor / Voc_cold)` | ขีดจำกัดออกแบบที่มี margin |
| Nmin MPPT | `CEILING(MPPT min voltage / Vmp_hot)` | จำนวนแผงขั้นต่ำ/string |
| Cable resistance | `R = rho × temperature factor × loop length / area + connector allowance` | ใช้ loop length ไม่ใช่ one-way |
| Voltage drop (V) | `ΔV = I × R` | แรงดันตกคร่อมสายหน่วย V |
| Voltage drop (%) | `ΔV / String Vmp × 100` | ≤ เกณฑ์ที่ตั้ง (%) |
| Power loss | `I² × R / String power × 100` | ≤ เกณฑ์ที่ตั้ง (%) |
""")
    st.subheader("ตรวจสอบความต้านทานและ Voltage Drop สาย DC")
    st.caption(
        "R_total = ρ × Temperature factor × Loop length ÷ Area "
        "+ Connector allowance • ΔV = Imp × R_total • "
        "Voltage drop (%) = ΔV ÷ String Vmp × 100"
    )
    if design["cables"].empty:
        st.info("ยังไม่มีข้อมูลสาย DC สำหรับตรวจสอบ")
    else:
        cable_resistance_check = design["cables"][[
            "string_id", "inverter_id", "one_way_m", "loop_m", "material",
            "resistivity_ohm_mm2_m", "temperature_factor", "size_mm2",
            "conductor_resistance_ohm", "connector_allowance_ohm",
            "resistance_ohm", "imp_a", "string_vmp_v",
            "voltage_drop_v", "voltage_drop_pct",
            "voltage_drop_limit_pct", "voltage_drop_status",
        ]].copy()
        cable_resistance_check = cable_resistance_check.rename(columns={
            "one_way_m": "One-way (m)",
            "loop_m": "Loop (m)",
            "resistivity_ohm_mm2_m": "ρ (Ω·mm²/m)",
            "temperature_factor": "Temp. factor",
            "size_mm2": "Area (mm²)",
            "conductor_resistance_ohm": "Conductor R (Ω)",
            "connector_allowance_ohm": "Connector R (Ω)",
            "resistance_ohm": "Total R (Ω)",
            "imp_a": "Imp (A)",
            "string_vmp_v": "String Vmp (V)",
            "voltage_drop_v": "Voltage drop (V)",
            "voltage_drop_pct": "Voltage drop (%)",
            "voltage_drop_limit_pct": "Limit (%)",
            "voltage_drop_status": "VD check",
        })
        resistance_styler = cable_resistance_check.style.format(
            {
                "One-way (m)": "{:,.2f}",
                "Loop (m)": "{:,.2f}",
                "ρ (Ω·mm²/m)": "{:.6f}",
                "Temp. factor": "{:.3f}",
                "Area (mm²)": "{:,.2f}",
                "Conductor R (Ω)": "{:.6f}",
                "Connector R (Ω)": "{:.6f}",
                "Total R (Ω)": "{:.6f}",
                "Imp (A)": "{:.3f}",
                "String Vmp (V)": "{:.3f}",
                "Voltage drop (V)": "{:.3f}",
                "Voltage drop (%)": "{:.3f}",
                "Limit (%)": "{:.3f}",
            },
            na_rep="-",
        ).map(status_style, subset=["VD check"])
        st.dataframe(resistance_styler, width="stretch", hide_index=True)
    if not design.get("critical_missing"):
        st.json({"current_inputs": {"Voc cold V": design["limits"]["voc_cold_v"], "Vmp hot V": design["limits"]["vmp_hot_v"],
                                           "Nmin MPPT": design["limits"]["nmin_mppt"], "Nmax design": design["limits"]["nmax_design"],
                                           "Cable material": cable_material,
                                           "Cable resistivity Ω·mm²/m": 0.0175 if cable_material == "Copper" else 0.0282,
                                           "Cable temperature factor": 1.2,
                                           "Connector allowance Ω": 0.002,
                                           "Cable mm2": cable_size}})

with tab6:
    st.subheader("คู่มือใช้งานแบบย่อ")
    st.markdown("""
1. ไปที่ **Master Data** และตรวจ/แก้ข้อมูล datasheet ของแผงและ inverter ให้ตรงรุ่น/market/revision
2. ช่องสำคัญที่ต้องตรวจสอบคือ **รุ่น Inverter** หน้า 1, **จำนวนแผงรวม** และ **Scope DC max V** หน้า 2 (ช่องอยู่ในกรอบเหลือง)
3. เลือก **รุ่น Inverter** เป็น `AUTO` เมื่อต้องให้ระบบเลือกจำนวนเครื่องน้อยที่สุดและจัด String/MPPT ได้ครบ
4. ใน **Auto-layout & String** กรอก **จำนวนแผงรวม** เป็นเลขคู่ เพราะ Rapid Shutdown / Optimizer อัตรา 2:1 และจำนวนแผงระหว่าง String ต่างกันไม่เกิน 2 แผง
5. เลือก **Scope DC max V** เพื่อกำหนด Nmax ด้านแรงดัน PV; `AUTO` จะอ้างอิง Inverter จากหน้า 1 และตัวเลือกจะแสดงเฉพาะค่าที่มีใน Master Inverter
6. ถ้ามีเพียงจำนวนแผง ให้กด **ใช้ Auto-layout** เพื่อสร้าง Candidate strings และนำไปจัด MPPT
7. ถ้ามี layout อยู่แล้ว ให้กรอกแต่ละ roof group: จำนวนแผง, orientation, shading และ one-way cable route
8. ดู **MPPT & Cable**: สถานะเขียว = PASS, แดง = FAIL, เหลือง = ต้องทบทวน พร้อม comment แนบท้ายทุกจุดที่ไม่ผ่าน
9. ดู **QA/QC & PVsyst** ก่อนดาวน์โหลด CSV/JSON และต้องเพิ่มไฟล์ PAN/OND ที่ผ่านการยืนยันก่อนใช้ PVsyst
10. ตาราง **Export Excel — String / MPPT** และ **Export Excel — Input + Assignment (16 คอลัมน์)** ด้านล่างหน้า 1 ใช้ตรวจสอบและกดปุ่ม Export เพื่อดาวน์โหลด; ตารางรวม 16 คอลัมน์เรียงข้อมูลกรอก 1–11 และผลจัด 12–16 ตามรูป
""")
    st.info("ไม่ควรแก้ข้อมูล master โดยไม่มี datasheet source; ห้ามถือว่า PASS หากช่อง verification ยังไม่ Verified")

with tab7:
    st.subheader("Editable Datasheet Master")
    st.caption("แก้ค่าไฟฟ้าได้โดยตรง; ผลทุกหน้าจะคำนวณใหม่ทันทีในรอบถัดไป โปรดเปลี่ยน Verification status และ Source ให้ครบ")
    st.session_state.module_master = st.data_editor(master_modules, num_rows="dynamic", use_container_width=True, key="module_master_editor")
    st.session_state.inverter_master = st.data_editor(master_inverters, num_rows="dynamic", use_container_width=True, key="inverter_master_editor")
